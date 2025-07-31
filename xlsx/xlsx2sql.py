import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
# ROLE_SYS,시스템 관리자
# ROLE_PTWKR,중대시민재해 총괄 담당자
# ROLE_PUWKR,중대시민재해 실국본부 담당자
# ROLE_PDWKR,중대시민재해 부서 담당자
# ROLE_PBWKR,중대시민재해 업무 담당자
# ROLE_ITWKR,중대산업재해 총괄 담당자
# ROLE_IUWKR,중대산업재해 실국본부 담당자
# ROLE_IDWKR,중대산업재해 부서 담당자
# ROLE_IBWKR,중대산업재해 업무 담당자
# ROLE_EWKR,안전교육 담당자
# ROLE_USER,일반 사용자

roles = {
    40: ("ROLE_SYS", "CREATE"),
    41: ("ROLE_SYS", "READ"),
    42: ("ROLE_SYS", "UPDATE"),
    43: ("ROLE_SYS", "DELETE"),
    44: ("ROLE_PTWKR", "CREATE"),
    45: ("ROLE_PTWKR", "READ"),
    46: ("ROLE_PTWKR", "UPDATE"),
    47: ("ROLE_PTWKR", "DELETE"),
    48: ("ROLE_PUWKR", "CREATE"),
    49: ("ROLE_PUWKR", "READ"),
    50: ("ROLE_PUWKR", "UPDATE"),
    51: ("ROLE_PUWKR", "DELETE"),
    52: ("ROLE_PDWKR", "CREATE"),
    53: ("ROLE_PDWKR", "READ"),
    54: ("ROLE_PDWKR", "UPDATE"),
    55: ("ROLE_PDWKR", "DELETE"),
    56: ("ROLE_PBWKR", "CREATE"),
    57: ("ROLE_PBWKR", "READ"),
    58: ("ROLE_PBWKR", "UPDATE"),
    59: ("ROLE_PBWKR", "DELETE"),
    60: ("ROLE_ITWKR", "CREATE"),
    61: ("ROLE_ITWKR", "READ"),
    62: ("ROLE_ITWKR", "UPDATE"),
    63: ("ROLE_ITWKR", "DELETE"),
    64: ("ROLE_IUWKR", "CREATE"),
    65: ("ROLE_IUWKR", "READ"),
    66: ("ROLE_IUWKR", "UPDATE"),
    67: ("ROLE_IUWKR", "DELETE"),
    68: ("ROLE_IDWKR", "CREATE"),
    69: ("ROLE_IDWKR", "READ"),
    70: ("ROLE_IDWKR", "UPDATE"),
    71: ("ROLE_IDWKR", "DELETE"),
    72: ("ROLE_IBWKR", "CREATE"),
    73: ("ROLE_IBWKR", "READ"),
    74: ("ROLE_IBWKR", "UPDATE"),
    75: ("ROLE_IBWKR", "DELETE"),
    76: ("ROLE_EWKR", "CREATE"),
    77: ("ROLE_EWKR", "READ"),
    78: ("ROLE_EWKR", "UPDATE"),
    79: ("ROLE_EWKR", "DELETE"),
    80: ("ROLE_USER", "CREATE"),
    81: ("ROLE_USER", "READ"),
    82: ("ROLE_USER", "UPDATE"),
    83: ("ROLE_USER", "DELETE"),
}

def main():
    wb = openpyxl.load_workbook('Res.xlsx')
    ws = wb.active

    rows = ws.iter_rows(min_row=12, max_row=ws.max_row, values_only=True)
    prev_top_menu = None
    for row in rows:
        if row[39] is None:
            continue
        top_menu = row[0] if row[0] is not None else prev_top_menu
        prev_top_menu = top_menu
        if row[39] is None:
            continue
        for idx, (role, code) in roles.items():
            # if role != "ROLE_USER":
            #     continue
            if row[idx] is None:
                continue
            col_nm = get_column_letter(idx + 1)
            # print(f"Top Menu Name: {prev_top_menu}, Menu Name: {row[1]}, Menu No: {row[39]}, Role: {role}, Code: {code}, Column: {col_nm}")
            print(f"insert into comtnmenucreatdtls (menu_no, author_code, mapng_creat_id, reg_dt, allow_yn, auth_type) "
                  f"values ('{row[39]}', '{role}', null, '20250728000000', 'Y', '{code}');")


        # if row[40] is None:
        #     continue
        # menu_no = row[39]
        # for idx, cell in enumerate(row, start=1):
        #     col_nm = get_column_letter(idx)
        #     if col_nm == "B":
        #         print(f"Menu Name: {row[idx - 1]}, Menu No: {menu_no}")

if __name__ == "__main__":
    main()
